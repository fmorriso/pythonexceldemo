import os
import sys
from importlib.metadata import version

import openpyxl

from logging_utility import LoggingUtility as LU


def get_required_package_names() -> list[str]:
    """
    read the requirements.txt file and return a sorted list of package names.
    :return: sorted list of package names
    :rtype: list[str
    """
    packages: list[str] = []
    with open('requirements.txt') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue  # skip blank lines and comments
            package = line.split('~')[0].strip()  # works for ~=, >=, ==, etc.
            packages.append(package)

    packages.sort(key = str.lower)
    return packages


def get_python_version() -> str:
    return f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"


def get_package_version(package_name: str) -> str:
    return version(package_name)


def read_excel_workbook():
    LU.debug("inside readWorkbook")
    LU.debug(f"current working directory is: {os.getcwd()}")
    # C:\Users\frede\Dropbox\Family Room\S-Drive\Holiday Cards\2022 Card list.xlsx
    # C:\Users\frede\Dropbox\download\python\Automate_the_Boring_Stuff\example.xlsx
    # fn : str = r"C:\Users\frede\Dropbox\Family Room\S-Drive\Holiday Cards\2022 Card list.xlsx"
    # fn: str = r"C:\Users\frede\Dropbox\download\python\Automate_the_Boring_Stuff\example.xlsx"
    fn: str = "example.xlsx"
    wb = openpyxl.load_workbook(filename = fn, read_only = True)
    LU.debug(f"wb typs: {type(wb)}")
    LU.debug("after load_workbook")
    LU.debug(f"sheetnames: {wb.sheetnames}")
    ws = wb.active
    print(f"The active worksheet title={ws.title}")


if __name__ == '__main__':
    LU.start_logging('log.txt')

    msg = f'Python version: {get_python_version()}'
    LU.log_info_and_debug(msg)

    package_names = get_required_package_names()

    for pkg in package_names:
        package_name = f'{pkg}'.ljust(16)
        try:
            LU.log_info_and_debug(f'{package_name}{get_package_version(pkg)}')
        except Exception as e:
            print(e)

    read_excel_workbook()
